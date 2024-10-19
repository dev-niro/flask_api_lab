from flask import Flask, jsonify, request
import openpyxl
from flask_swagger_ui import get_swaggerui_blueprint

app = Flask(__name__)

# Swagger setup
SWAGGER_URL = '/swagger'  # URL for exposing Swagger UI
API_URL = '/static/swagger.yaml'  # Path to the Swagger YAML file

swaggerui_blueprint = get_swaggerui_blueprint(SWAGGER_URL, API_URL, config={'app_name': "Example API"})
app.register_blueprint(swaggerui_blueprint, url_prefix=SWAGGER_URL)

# Mensaje de bienvenida
@app.route('/')
def welcome():
    return "Welcome to my Flask API!"

# Ruta para obtener todos los datos del archivo Excel (GET /usuarios)
@app.route('/usuarios', methods=['GET'])
def obtener_usuarios():
    usuarios = []
    libro = openpyxl.load_workbook('datos.xlsx')
    hoja = libro.active
    for fila in hoja.iter_rows(values_only=True):
        id, nombre, email = fila
        usuarios.append({'id': id, 'nombre': nombre, 'email': email})
    return jsonify(usuarios)

# Ruta para agregar un nuevo usuario al archivo Excel (POST /usuarios)
@app.route('/usuarios', methods=['POST'])
def agregar_usuario():
    nuevo_usuario = request.json
    libro = openpyxl.load_workbook('datos.xlsx')
    hoja = libro.active
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        id_existente = str(fila[0])
        if id_existente == str(nuevo_usuario['id']):
            return jsonify({'mensaje': 'Error: El usuario con ID ya existe'}), 400
    hoja.append([nuevo_usuario['id'], nuevo_usuario['nombre'], nuevo_usuario['email']])
    libro.save('datos.xlsx')
    return jsonify({'mensaje': 'Usuario agregado'}), 201

# Ruta para obtener un usuario específico por ID (GET /usuarios/<id>)
@app.route('/usuarios/<id>', methods=['GET'])
def obtener_usuario(id):
    libro = openpyxl.load_workbook('datos.xlsx')
    hoja = libro.active
    for fila in hoja.iter_rows(values_only=True):
        id_usuario, nombre, email = fila
        if str(id_usuario) == id:
            return jsonify({'id': id_usuario, 'nombre': nombre, 'email': email})
    return jsonify({'mensaje': 'Usuario no encontrado'}), 404

# Ruta para actualizar un usuario (PUT /usuarios/<id>)
@app.route('/usuarios/<id>', methods=['PUT'])
def actualizar_usuario(id):
    datos_actualizados = request.json
    libro = openpyxl.load_workbook('datos.xlsx')
    hoja = libro.active
    actualizado = False
    for fila in hoja.iter_rows(min_row=2, values_only=False):
        id_usuario = fila[0].value
        if str(id_usuario) == id:
            fila[1].value = datos_actualizados.get('nombre', fila[1].value)
            fila[2].value = datos_actualizados.get('email', fila[2].value)
            actualizado = True
            break
    if actualizado:
        libro.save('datos.xlsx')
        return jsonify({'mensaje': 'Usuario actualizado'}), 200
    return jsonify({'mensaje': 'Usuario no encontrado'}), 404

# Ruta para eliminar un usuario (DELETE /usuarios/<id>)
@app.route('/usuarios/<id>', methods=['DELETE'])
def borrar_usuario(id):
    libro = openpyxl.load_workbook('datos.xlsx')
    hoja = libro.active
    actualizado = False
    for fila in hoja.iter_rows(min_row=2, values_only=False):
        id_usuario = fila[0].value
        if str(id_usuario) == id:
            hoja.delete_rows(fila[0].row)
            actualizado = True
            break
    if actualizado:
        libro.save('datos.xlsx')
        return jsonify({'mensaje': 'Usuario borrado'}), 200
    return jsonify({'mensaje': 'Usuario no encontrado'}), 404

if __name__ == '__main__':
    app.run(debug=True)
